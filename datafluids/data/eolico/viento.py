import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import statistics

st.title('CALCULO DE LA POTENCIA GENERADA POR UN AEROGENERADOR BASADO EN LA VELOCIDAD DEL VIENTO DE VILLAVICENCIO')
archivo ='C:/Users/juan/Documents/python/viento/horaldos.xlsx'
archivodos='C:/Users/juan/Documents/python/viento/mensualdos.xlsx'
archivotres='C:/Users/juan/Documents/python/viento/generador.xlsx'
excel_document = openpyxl.load_workbook('C:/Users/juan/Documents/python/viento/horaldos.xlsx')
excel_documentdos= openpyxl.load_workbook('C:/Users/juan/Documents/python/viento/mensualdos.xlsx')

horal = pd.read_excel(archivo, sheet_name='Table 2')
horaldos=excel_document.get_sheet_by_name('Table 2')
mensualdos=excel_documentdos.get_sheet_by_name('Table 2')
mensual= pd.read_excel(archivodos, sheet_name='Table 2')
generador= pd.read_excel(archivotres, sheet_name='Hoja1')

horal.describe()
mensual.describe()
generador.describe()
st.header('VELOCIDAD DEL VIENTO A LO LARGO DE UN DIA')
col1, col2= st.beta_columns(2)
with col1:
    st.dataframe(horal)
with col2:
    st.line_chart(horal)
st.header('VELOCIDAD DEL VIENTO EN UN AÑO')
col1, col2=st.beta_columns(2)
with col1:
    st.dataframe(mensual)
with col2:
    st.line_chart(mensual)
st.header('CURVA DE POTENCIA DE UN AEROGENERADOR')
col1, col2=st.beta_columns(2)
with col1:
    st.dataframe(generador)
with col2:
    st.line_chart(generador)

x=0;
y=0;
listaHora=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23]
potenciaDiaria=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23]
while x<23:
    listaHora[x]=horaldos.cell(row = x+2, column = 1).value
    x+=1

while y<23:
    if listaHora[y]<2:
        potenciaDiaria[y]=listaHora[y]*0.6
    else:
        potenciaDiaria[y]=-0.00045*(listaHora[y])**5+0.03178*(listaHora[y])**4-0.8144*(listaHora[y])**3+8.48259*(listaHora[y])**2-23.58802*(listaHora[y])+20.50708
    y=y+1
mean = statistics.mean(potenciaDiaria)
st.header('POTENCIA ENTREGADA EN UN DIA')
st.line_chart(potenciaDiaria)
st.header('Potencia promedio diaria')
st.write(mean)
x=0;
y=0;
listames=[1,2,3,4,5,6,7,8,9,10,11,12]
potenciames=[1,2,3,4,5,6,7,8,9,10,11,12]
while x<12:
    listames[x]=mensualdos.cell(row = x+2, column = 1).value
    x+=1


while y<12:
    if listames[y]<2:
        potenciames[y]=listames[y]*0.6
    else:
        potenciames[y]=-0.00045*(listames[y])**5+0.03178*(listames[y])**4-0.8144*(listames[y])**3+8.48259*(listames[y])**2-23.58802*(listames[y])+20.50708
    y=y+1
st.header('POTENCIA GENERADA EN UN AÑO')
st.line_chart(potenciames)
meandos = statistics.mean(potenciames)
st.header('Potencia promedio anual')
st.write(meandos)



